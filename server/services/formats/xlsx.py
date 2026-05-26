"""XLSX 解析 — 逐 sheet 转表格文本。"""

from pathlib import Path


def parse_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        parts.append(f"## {name}")
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            parts.append("\n".join(rows))
        parts.append("")
    wb.close()
    return "\n".join(parts)

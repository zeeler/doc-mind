import pytest
from pathlib import Path


class TestXlsxParser:
    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws['A1'] = "Name"
            ws['B1'] = "Value"
            ws['A2'] = "Revenue"
            ws['B2'] = "1000"
            path = tmp_path / "test.xlsx"
            wb.save(str(path))
            return path
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_parse_xlsx(self, sample_xlsx):
        from server.services.formats.xlsx import parse_xlsx
        text = parse_xlsx(sample_xlsx)
        assert "Sheet1" in text
        assert "Revenue" in text
        assert "1000" in text

    def test_parse_empty_xlsx(self, tmp_path):
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            path = tmp_path / "empty.xlsx"
            wb.save(str(path))
            from server.services.formats.xlsx import parse_xlsx
            text = parse_xlsx(path)
            assert isinstance(text, str)
        except ImportError:
            pytest.skip("openpyxl not installed")
